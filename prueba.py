import openpyxl
import random
from openpyxl import Workbook
from datetime import datetime, timedelta

# Crea un nuevo libro de Excel
workbook = Workbook()

# Crea una nueva hoja en el libro de Excel
sheet = workbook.active

# Define los nombres de las neveras y congeladores
neveras = ["Nevera 1", "Nevera 2", "Nevera 3", "Nevera 4", "Nevera 5", "Nevera 6"]
congeladores = ["Congelador 1", "Congelador 2", "Congelador 3"]

# Escribe los encabezados de la tabla
sheet.cell(row=1, column=1, value="Fecha")
for i, nevera in enumerate(neveras + congeladores, start=2):
    sheet.cell(row=1, column=i, value=nevera)

# Define la fecha de inicio y la fecha actual
fecha_inicio = datetime(2023, 1, 1)
fecha_actual = datetime(2023, 6, 30)

# Itera sobre las fechas y las neveras y escribe los datos en la tabla
fila = 2
fecha = fecha_inicio
while fecha <= fecha_actual:
    sheet.cell(row=fila, column=1, value=fecha.strftime("%Y-%m-%d"))
    for j, nevera in enumerate(neveras + congeladores, start=2):
        # Genera un valor de temperatura aleatorio entre 2°C y 8°C para las neveras
        # y entre -18°C y -12°C para los congeladores
        if j <= 5:
            temp = round(random.uniform(3, 8), 1)
        else:
            temp = round(random.uniform(-18, -12), 1)
        sheet.cell(row=fila, column=j, value=temp)
    fecha += timedelta(days=3.5)
    fila += 1

# Guarda el libro de Excel
try:
    workbook.save("temperaturas2023.xlsx")
    print('archivo guardado')
except:
    print('Ha ocurrido algún error')
